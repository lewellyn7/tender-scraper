#!/usr/bin/env python3
"""Regenerate latest.json from latest Excel backup"""
import sys, json
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
import openpyxl

EXCEL_FILE = Path("/home/lewellyn/tender-scraper/output/chongqing_tender_v3_20260429_133313.xlsx")
OUTPUT_FILE = Path("/home/lewellyn/tender-scraper/output/latest.json")

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb.active

headers = [cell.value for cell in ws[1]]
print(f"Headers ({len(headers)}): {headers[:10]}...")

projects = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    proj = dict(zip(headers, row))
    projects.append(proj)

print(f"Rows: {len(projects)}")

output_data = {
    "total": len(projects),
    "filtered": len([p for p in projects if p.get("keywords_matched")]),
    "last_run": "2026-04-29 13:33:13",
    "projects": projects,
}

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"Written: {OUTPUT_FILE}")
print(f"Total: {len(projects)}, filtered: {output_data['filtered']}")
