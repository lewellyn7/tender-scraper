"""修复 xlsx 文件里残留的 _1 错版 URL (2026-07-22)

背景 (用户拍板 2026-07-22 13:17):
  12:01 中午汇报引用的 xlsx 是 10:17 collector 跑旧代码时生成的快照,
  里面 74/311 URL 带 _1 (全 014005001). 用户中午看的还是错版链接.
  13:00 SQL UPDATE 只改了 DB, 没反向更新 xlsx 文件. 必须从干净 DB 重写 xlsx.

策略:
  1. SELECT id, url, title, publish_date FROM projects_cqggzy
     (按 trade_id + infoid + categoryNum 建索引, 同一个 URL 可能有多个 id, 取 publish_date 最近)
  2. openpyxl load xlsx
  3. 对每行第16列「链接」:
     - extract infoid (剥 _N) + categoryNum
     - 在 dict 里查 DB 干净 URL
     - 如果命中 → 用 DB URL 覆盖
     - 没命中 → 保留原 URL (可能是 UUID 格式, 不在本次 bug 修复范围)
  4. backup 原文件 (output/chongqing_tender_v3_20260722_101757.xlsx.bak-20260722_13XX)
  5. save 到原文件

用法:
  python scripts/fix_xlsx_url_suffix_2026-07-22.py [xlsx_path]
  默认: output/chongqing_tender_v3_20260722_101757.xlsx

依赖: pip install openpyxl psycopg2-binary python-dotenv
"""
from __future__ import annotations

import os
import re
import sys
import shutil
from datetime import datetime
from typing import Dict, Tuple, Optional

import openpyxl

# 路径 (默认 vs 命令行)
DEFAULT_XLSX = "output/chongqing_tender_v3_20260722_101757.xlsx"
URL_COL = 15  # 0-indexed: 表头第16列「链接」
SOURCE_URL_COL = 16  # 第17列「来源页」

# DB 连接 (复用 .env 里的配置, 默认值与 .env 保持一致)
# .env: DB_HOST=localhost, DB_NAME=tender_scraper, DB_USER=root, DB_PASSWORD=root123
DB_ENV = {
    "POSTGRES_HOST": os.getenv("DB_HOST") or os.getenv("POSTGRES_HOST", "localhost"),
    "POSTGRES_PORT": os.getenv("DB_PORT") or os.getenv("POSTGRES_PORT", "5432"),
    "POSTGRES_USER": os.getenv("DB_USER") or os.getenv("POSTGRES_USER", "root"),
    "POSTGRES_PASSWORD": os.getenv("DB_PASSWORD") or os.getenv("POSTGRES_PASSWORD", "root123"),
    "POSTGRES_DB": os.getenv("DB_NAME") or os.getenv("POSTGRES_DB", "tender_scraper"),
}

# 加载 .env (路径默认是当前目录)
try:
    from dotenv import load_dotenv
    load_dotenv()
    # load 后重新读取
    DB_ENV = {
        "POSTGRES_HOST": os.getenv("DB_HOST") or os.getenv("POSTGRES_HOST", "localhost"),
        "POSTGRES_PORT": os.getenv("DB_PORT") or os.getenv("POSTGRES_PORT", "5432"),
        "POSTGRES_USER": os.getenv("DB_USER") or os.getenv("POSTGRES_USER", "root"),
        "POSTGRES_PASSWORD": os.getenv("DB_PASSWORD") or os.getenv("POSTGRES_PASSWORD", "root123"),
        "POSTGRES_DB": os.getenv("DB_NAME") or os.getenv("POSTGRES_DB", "tender_scraper"),
    }
except ImportError:
    pass  # 没装 dotenv 就用环境变量

URL_RE = re.compile(r"/trade/(\d+)/(\d+)(_\d+)?\?categoryNum=(\d+)")
# 不带 trade_id 的版本 (xlsx 里部分 URL 可能简写)
URL_RE_SHORT = re.compile(r"/(\d+)(_\d+)?\?categoryNum=(\d+)")


def _connect():
    """用环境变量直连 PostgreSQL (避免依赖 app.utils.db 路径)"""
    import psycopg2
    return psycopg2.connect(
        host=DB_ENV["POSTGRES_HOST"],
        port=int(DB_ENV["POSTGRES_PORT"]),
        user=DB_ENV["POSTGRES_USER"],
        password=DB_ENV["POSTGRES_PASSWORD"],
        database=DB_ENV["POSTGRES_DB"],
    )


def _extract_url_key(url: str) -> Optional[Tuple[str, str]]:
    """从 URL extract (trade_id, infoid, categoryNum) 三元组. 返回 (infoid, categoryNum) for indexing."""
    if not url:
        return None
    m = URL_RE.search(url)
    if m:
        # group 2 是 infoid (剥 _N), group 4 是 categoryNum
        return (m.group(2), m.group(4))
    # 短 URL (UUID 格式, 不剥 _N 因为本来就没)
    m = URL_RE_SHORT.search(url)
    if m and not m.group(1).isdigit():
        # UUID 格式, 整体作为 infoid
        return (m.group(1), m.group(3))
    return None


def _load_db_url_index() -> Dict[Tuple[str, str], str]:
    """从 DB 加载 (infoid, categoryNum) -> 干净 URL 的索引

    只取最近 30 天的数据 (避免老数据干扰), 按 publish_date DESC 优先.
    """
    conn = _connect()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT url, publish_date FROM projects_cqggzy
            WHERE url IS NOT NULL
              AND url::text LIKE '%cqggzy.com%'
              AND (publish_date IS NULL OR publish_date >= (CURRENT_DATE - INTERVAL '30 days'))
            ORDER BY publish_date DESC NULLS LAST, id DESC
        """)
        index: Dict[Tuple[str, str], str] = {}
        for url, pd in cur.fetchall():
            if not url:
                continue
            key = _extract_url_key(url)
            if key and key not in index:
                # 同 key 保留第一条 (最近的)
                index[key] = url
        return index
    finally:
        conn.close()


def _normalize_url(url: str, db_url: str) -> str:
    """从 DB URL 提取标准化版本: 保持 BASE_URL/trade/{trade_id}/{infoid}?categoryNum={catnum}

    用 DB 的 trade_id + infoid + catnum, 不强制完整替换 prefix (避免破坏 https 等).
    """
    # 取 DB URL 的 infoid + catnum 部分
    m_db = URL_RE.search(db_url)
    if not m_db:
        return db_url
    db_trade_id = m_db.group(1)
    db_infoid = m_db.group(2)
    db_catnum = m_db.group(4)
    # 从原 URL 取 BASE (https://...trade/{trade_id}/{old_infoid}?categoryNum=...)
    # 替换 infoid 和 catnum
    return re.sub(
        r"(/trade/\d+/)\d+(_\d+)?(\?categoryNum=)\d+",
        rf"\g<1>{db_infoid}\g<3>{db_catnum}",
        url,
    )


def fix_xlsx(xlsx_path: str) -> Tuple[int, int]:
    """修复 xlsx 文件里残留的 _1 URL. 返回 (修复条数, 总带 _N 条数)."""
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"xlsx 文件不存在: {xlsx_path}")

    print(f"📂 加载 xlsx: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    print(f"   行数: {ws.max_row}, 列数: {ws.max_column}")

    print(f"🔌 连接 DB ({DB_ENV['POSTGRES_HOST']}:{DB_ENV['POSTGRES_PORT']}/{DB_ENV['POSTGRES_DB']})...")
    db_index = _load_db_url_index()
    print(f"   DB 索引: {len(db_index)} 条 (infoid, categoryNum) -> URL")

    fixed = 0
    total_with_n = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2):  # 跳过表头
        if len(row) <= URL_COL:
            continue
        cell = row[URL_COL]
        url = cell.value or ""
        if "_1" not in url and "_2" not in url and "_3" not in url:
            continue
        total_with_n += 1
        key = _extract_url_key(url)
        if not key:
            skipped += 1
            continue
        # 关键: 从 key 剥 _N 后查 DB
        # key 是 (infoid_with_N, catnum) - 但 _extract_url_key 已经返回 (剥后 infoid, catnum)
        # 所以直接查 db_index[key]
        db_url = db_index.get(key)
        if not db_url:
            skipped += 1
            continue
        new_url = _normalize_url(url, db_url)
        cell.value = new_url
        fixed += 1

    # 备份原文件
    backup_path = f"{xlsx_path}.bak-{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    shutil.copy2(xlsx_path, backup_path)
    print(f"💾 备份原文件: {backup_path}")

    # 保存 (覆盖)
    wb.save(xlsx_path)
    print(f"✅ 已保存: {xlsx_path}")
    print(f"📊 统计: 总带 _N={total_with_n}, 已修复={fixed}, 跳过={skipped} (DB 无匹配)")

    return fixed, total_with_n


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    fixed, total = fix_xlsx(xlsx_path)
    print(f"\n🎯 完成: 修复 {fixed}/{total} 条 URL")
    if fixed < total:
        print(f"⚠️  剩余 {total - fixed} 条带 _N URL 在 DB 无匹配 (可能是 UUID 或新发布的 URL)")
    sys.exit(0)


if __name__ == "__main__":
    main()